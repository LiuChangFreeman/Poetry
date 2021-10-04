#encoding=utf-8
import re
import os
import openpyxl
from openpyxl.styles import Alignment

def main():
    path_current = os.path.split(os.path.realpath(__file__))[0]
    regex=re.compile(r"重.*紫")
    visited=set()
    results=[]
    for file_csv in os.listdir(path_current):
        if not file_csv.endswith(".csv"):
            continue
        for line in open(os.path.join(path_current,file_csv)).readlines():
            try:
                title,dynasty,author,content=line.split(",")
            except:
                continue
            title=title.strip("\"")
            dynasty=dynasty.strip("\"")
            author=author.strip("\"")
            words=content.strip("\"").split("。")
            for word in words:
                if word in visited:
                    continue
                match=re.findall(regex,word)
                if len(match)!=0:
                    visited.add(word)
                    results.append((title,dynasty,author,word))

    wb = openpyxl.Workbook()
    center = Alignment(horizontal='center', vertical='center')
    sheet = wb.create_sheet(u"统计", 0)
    titles = ["题目","朝代","作者","内容片段"]
    sheet.column_dimensions['A'].width = 24.0
    sheet.column_dimensions['B'].width = 12.0
    sheet.column_dimensions['C'].width = 12.0
    sheet.column_dimensions['D'].width = 80.0
    for i in range(len(titles)):
        sheet.cell(row=1, column=i + 1).value = titles[i]
        sheet.cell(row=1, column=i + 1).alignment = center
    count = 2
    for item in results:
        sheet.cell(row=count, column=1).value = item[0]
        sheet.cell(row=count, column=2).value = item[1]
        sheet.cell(row=count, column=3).value = item[2]
        sheet.cell(row=count, column=4).value = item[3]
        sheet.cell(row=count, column=1).alignment = center
        sheet.cell(row=count, column=2).alignment = center
        sheet.cell(row=count, column=3).alignment = center
        sheet.cell(row=count, column=4).alignment = center
        count += 1
    wb.save(u"result.xlsx")

if __name__=="__main__":
    main()